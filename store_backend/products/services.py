from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import re
import time

import logging
import requests
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import BundleItem, MTGCard, PricingSettings, Product, PurchaseOrder, PurchaseOrderItem, SealedProduct, SingleCard, Supplier

SCRYFALL_BASE = "https://api.scryfall.com"
SCRYFALL_TIMEOUT = 10
logger = logging.getLogger(__name__)


COLUMN_ALIASES = {
    "name": ["name", "nombre"],
    "type": ["type", "tipo"],
    "price_clp": ["price_clp", "precio", "price"],
}

CATALOG_REQUIRED_HEADERS = ["name", "type", "price_clp"]
SINGLE_PURCHASE_REQUIRED_HEADERS = ["name", "condition", "qty", "price_usd", "total_usd", "foil"]


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _resolve_catalog_headers(raw_headers):
    normalized_headers = [_normalize_header(h) for h in raw_headers]
    logger.info("Headers recibidos en importación catálogo XLSX: %s", normalized_headers)

    alias_to_canonical = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_to_canonical[_normalize_header(alias)] = canonical

    header_map = {}
    for header in normalized_headers:
        canonical = alias_to_canonical.get(header, header)
        if canonical not in header_map:
            header_map[canonical] = header

    expected = CATALOG_REQUIRED_HEADERS
    missing = [column for column in expected if column not in header_map]
    if missing:
        raise ValidationError({
            "detail": "Columnas inválidas",
            "expected": expected,
            "received": normalized_headers,
        })

    return normalized_headers, header_map



def _detect_xlsx_format(normalized_headers):
    header_set = set(normalized_headers)
    if set(CATALOG_REQUIRED_HEADERS).issubset(header_set):
        return "catalog"
    if set(SINGLE_PURCHASE_REQUIRED_HEADERS).issubset(header_set):
        return "single_purchase_items"
    raise ValidationError({
        "detail": "Formato XLSX no reconocido",
        "received_headers": normalized_headers,
        "valid_formats": {
            "catalog": CATALOG_REQUIRED_HEADERS,
            "single_purchase_items": SINGLE_PURCHASE_REQUIRED_HEADERS,
        },
    })


def import_single_purchase_catalog_row(row_data):
    card, card_data, _ = resolve_scryfall_card(name=row_data.get("name"))
    product, created = Product.objects.update_or_create(
        name=str(row_data.get("name") or card.name).strip(),
        product_type=Product.ProductType.SINGLE,
        defaults={"price_clp": 0, "description": card.type_line or "", "image": card.image_large or card.image_normal or card.image_small or "", "is_active": True},
    )
    SingleCard.objects.update_or_create(
        product=product,
        defaults={
            "mtg_card": card,
            "condition": str(row_data.get("condition") or Product.CardCondition.NM).upper(),
            "is_foil": _to_bool(row_data.get("foil"), False),
            "edition": card.set_name,
            "price_usd_reference": _to_decimal(row_data.get("price_usd"), Decimal("0")) or extract_usd_price(card_data, _to_bool(row_data.get("foil"), False)),
        },
    )
    return product, created, []

class ScryfallServiceError(Exception):
    pass

# ... keep existing helpers

def _request_json(path, params=None):
    url = f"{SCRYFALL_BASE}{path}"
    try:
        response = requests.get(url, params=params, timeout=SCRYFALL_TIMEOUT)
    except requests.RequestException as exc:
        raise ScryfallServiceError("Error de red consultando Scryfall") from exc
    if response.status_code == 404:
        raise ScryfallServiceError("Carta no encontrada en Scryfall")
    if response.status_code != 200:
        raise ScryfallServiceError(f"Error Scryfall ({response.status_code})")
    try:
        return response.json()
    except ValueError as exc:
        raise ScryfallServiceError("Respuesta inválida desde Scryfall") from exc

def _image_uris(card_data):
    image_uris = card_data.get("image_uris") or {}
    if image_uris:
        return image_uris
    for face in card_data.get("card_faces") or []:
        if face.get("image_uris"):
            return face["image_uris"]
    return {}

def _normalize_card_data(card_data):
    image_uris = _image_uris(card_data)
    released = card_data.get("released_at")
    return {"name": card_data.get("name", ""), "printed_name": card_data.get("printed_name", ""), "set_name": card_data.get("set_name", ""), "set_code": card_data.get("set", ""), "collector_number": card_data.get("collector_number", ""), "rarity": card_data.get("rarity", ""), "mana_cost": card_data.get("mana_cost", ""), "type_line": card_data.get("type_line", ""), "oracle_text": card_data.get("oracle_text", ""), "colors": card_data.get("colors") or [], "color_identity": card_data.get("color_identity") or [], "image_small": image_uris.get("small", ""), "image_normal": image_uris.get("normal", ""), "image_large": image_uris.get("large", ""), "scryfall_uri": card_data.get("scryfall_uri", ""), "released_at": date.fromisoformat(released) if released else None, "raw_data": card_data}

def _to_decimal(value, fallback=Decimal("0")):
    try: return Decimal(str(value))
    except (InvalidOperation, TypeError): return fallback

def _to_int(value, default=0):
    try:
        if value in (None, ""):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        raise ValidationError("Valor numérico inválido")

def _to_bool(value, default=False):
    if value is None: return default
    if isinstance(value, bool): return value
    return str(value).strip().lower() in {"true","1","yes","on","si"}

def get_active_pricing_settings():
    settings = PricingSettings.objects.filter(is_active=True).order_by("-updated_at").first()
    if settings: return settings
    return PricingSettings(usd_to_clp=Decimal("1000"), import_factor=Decimal("1.30"), risk_factor=Decimal("1.10"), margin_factor=Decimal("1.25"), rounding_to=100)

def extract_usd_price(card_data, is_foil=False):
    prices = card_data.get("prices") or {}
    return _to_decimal(prices.get("usd_foil" if is_foil else "usd"), fallback=Decimal("0"))

def search_cards(query):
    url = f"{SCRYFALL_BASE}/cards/search"
    try:
        response = requests.get(url, params={"q": query}, timeout=10)
    except requests.RequestException as exc:
        raise ScryfallServiceError("Error de red consultando Scryfall") from exc
    if response.status_code == 404:
        return []
    if response.status_code != 200:
        raise ScryfallServiceError(f"Error Scryfall ({response.status_code})")
    payload = response.json()
    return payload.get("data", [])
def get_card_by_id(scryfall_id): return _request_json(f"/cards/{scryfall_id}")
def get_scryfall_card_by_id(scryfall_id):
    try:
        return get_card_by_id(scryfall_id)
    except ScryfallServiceError as exc:
        logger.warning("Scryfall lookup failed for scryfall_id=%s error=%s", scryfall_id, exc)
        raise ValidationError(f"Scryfall no encontró la carta: {exc}") from exc

def import_card(scryfall_id):
    card_data = get_card_by_id(scryfall_id)
    card, _ = MTGCard.objects.update_or_create(scryfall_id=card_data["id"], defaults=_normalize_card_data(card_data))
    return card, card_data

def _normalize_card_name(name):
    normalized_name = " ".join(str(name or "").replace("\n", " ").split()).strip()
    if ":" in normalized_name:
        _, possible_name = normalized_name.split(":", 1)
        if possible_name.strip():
            normalized_name = possible_name.strip()
    return normalized_name


def _normalized_for_match(value):
    return " ".join(str(value or "").strip().lower().split())


def _pick_card_match(cards, normalized_name):
    if not cards:
        return None
    normalized_target = _normalized_for_match(normalized_name)
    for card_data in cards:
        if _normalized_for_match(card_data.get("name")) == normalized_target:
            return card_data
    return None


def resolve_scryfall_card(*, scryfall_id=None, name=None):
    if scryfall_id:
        card_data = get_card_by_id(str(scryfall_id).strip())
        card, _ = MTGCard.objects.update_or_create(scryfall_id=card_data["id"], defaults=_normalize_card_data(card_data))
        return card, card_data, []
    normalized_name = _normalize_card_name(name)
    if not normalized_name:
        raise ValidationError("name es obligatorio para single")
    attempted_queries = [f'!"{normalized_name}"', normalized_name]
    cards = []
    query_used = attempted_queries[0]
    for query in attempted_queries:
        query_used = query
        try:
            cards = search_cards(query)
        except ScryfallServiceError:
            continue
        if cards:
            break
    if not cards:
        raise ValidationError({
            "name": normalized_name,
            "error": "No se pudo resolver la carta en Scryfall",
            "query_used": query_used,
            "suggestion": "Agrega columna scryfall_id para importación exacta",
        })
    card_data = _pick_card_match(cards, normalized_name) or cards[0]
    if len(cards) > 1 and not _pick_card_match(cards, normalized_name):
        suggestions = [card.get("name") for card in cards[:5] if card.get("name")]
        raise ValidationError({
            "name": normalized_name,
            "error": "Resultado ambiguo en Scryfall",
            "query_used": query_used,
            "suggestions": suggestions,
            "suggestion": "Agrega columna scryfall_id para importación exacta",
        })
    card, _ = MTGCard.objects.update_or_create(scryfall_id=card_data["id"], defaults=_normalize_card_data(card_data))
    return card, card_data, ["single sin scryfall_id: se resolvió por nombre"]


VENDOR_CONDITION_MAP = {
    "NM": "NM", "MINT": "NM", "M": "NM",
    "EX": "LP", "EXCELLENT": "LP",
    "VG": "MP", "VERY GOOD": "MP",
    "G": "HP", "GOOD": "HP", "PLAYED": "HP",
    "PO": "DMG", "POOR": "DMG",
}


def parse_vendor_invoice_xlsx(excel_file):
    logger.info("Parsing vendor invoice xlsx")
    wb = load_workbook(excel_file, data_only=True)
    sheet = wb.active
    rows = [list(row or []) for row in sheet.iter_rows(values_only=True)]

    section_pattern = re.compile(r"^[A-Z\s]+(?:SINGLES|SEALED)$")
    parenthetical_pattern = re.compile(r"\s*\(([^)]+)\)\s*$")
    total_keys = {"subtotal": "subtotal_usd", "shipping": "shipping_usd", "sales tax": "tax_usd", "tax": "tax_usd", "total": "total_usd"}

    items, parse_warnings, sections_found = [], [], []
    totals = {"subtotal_usd": Decimal("0"), "shipping_usd": Decimal("0"), "tax_usd": Decimal("0"), "total_usd": Decimal("0")}

    sections = []
    for idx, row in enumerate(rows):
        c0 = str(row[0] if row else "" or "").strip()
        if c0 and section_pattern.match(c0):
            sections.append({"name": c0, "start_idx": idx, "header_idx": None})
            sections_found.append(c0)
    for i, section in enumerate(sections):
        search_end = sections[i + 1]["start_idx"] if i + 1 < len(sections) else len(rows)
        for idx in range(section["start_idx"] + 1, search_end):
            c0 = str(rows[idx][0] if rows[idx] else "" or "").strip().lower()
            if c0 == "description":
                section["header_idx"] = idx
                break

    for row_idx, row in enumerate(rows, start=1):
        c0 = str(row[0] if row else "" or "").strip()
        if not c0:
            continue
        key = c0.lower()
        if key in total_keys and len(row) > 4 and row[4] is not None:
            cleaned = re.sub(r"[^0-9.]", "", str(row[4]))
            if cleaned:
                totals[total_keys[key]] = Decimal(cleaned)

    for section_idx, section in enumerate(sections):
        if section["header_idx"] is None:
            parse_warnings.append(f"Sección '{section['name']}' sin headers Description/Style/Qty/Price/Total")
            continue
        section_end = sections[section_idx + 1]["start_idx"] if section_idx + 1 < len(sections) else len(rows)
        inferred_raw = section["name"].split()[0].strip().upper()
        inferred_condition = VENDOR_CONDITION_MAP.get(inferred_raw, "NM")
        for idx in range(section["header_idx"] + 1, section_end):
            row = rows[idx]
            description = str(row[0] if row else "" or "").strip()
            if not description or description.lower() == "description":
                continue
            low0 = description.lower()
            if low0 in total_keys or section_pattern.match(description):
                continue
            if ":" not in description:
                continue

            set_hint, card_name_raw = description.rsplit(":", 1)
            set_hint = set_hint.strip()
            card_name = card_name_raw.strip()
            variant_hint = ""
            variant_match = parenthetical_pattern.search(card_name)
            if variant_match:
                variant_hint = variant_match.group(1).strip()
                card_name = parenthetical_pattern.sub("", card_name).strip()

            style_raw = str(row[1] if len(row) > 1 and row[1] is not None else "").strip().upper()
            condition = VENDOR_CONDITION_MAP.get(style_raw, inferred_condition) if style_raw else inferred_condition
            try:
                qty = int(row[2])
            except (TypeError, ValueError):
                parse_warnings.append(f"Fila {idx + 1}: qty inválida '{row[2] if len(row) > 2 else None}'")
                continue
            try:
                price_usd = Decimal(str(row[3]))
            except (InvalidOperation, TypeError, ValueError):
                parse_warnings.append(f"Fila {idx + 1}: price inválido '{row[3] if len(row) > 3 else None}'")
                continue
            try:
                total_usd = Decimal(str(row[4])) if len(row) > 4 and row[4] is not None else qty * price_usd
            except (InvalidOperation, TypeError, ValueError):
                parse_warnings.append(f"Fila {idx + 1}: total inválido '{row[4] if len(row) > 4 else None}', usando qty*price")
                total_usd = qty * price_usd

            items.append({
                "row": idx + 1,
                "raw_description": description,
                "card_name": card_name,
                "set_hint": set_hint,
                "variant_hint": variant_hint,
                "is_foil": "foil" in set_hint.lower(),
                "condition": condition,
                "qty": qty,
                "price_usd": price_usd,
                "total_usd": total_usd,
            })

    return {"items": items, "totals": totals, "sections_found": sections_found, "parse_warnings": parse_warnings}


def resolve_scryfall_card_from_vendor(card_name, set_hint, is_foil):
    warnings = []
    cleaned = re.sub(r"\s+", " ", str(card_name or "").strip())
    m = re.search(r"\(([^)]+)\)", cleaned)
    variant_hint = m.group(1).strip() if m else ""
    cleaned = re.sub(r"\([^)]*\)", "", cleaned).strip()
    queries = [f'!"{cleaned}"']
    set_token = (set_hint or "").split()[0].lower() if set_hint else ""
    if set_token:
        queries.append(f'!"{cleaned}" {set_token}')
    queries.append(cleaned)
    all_cards = []
    for q in queries:
        try:
            time.sleep(0.1)
            cards = search_cards(q)
            if cards:
                all_cards = cards
                break
        except ScryfallServiceError as exc:
            warnings.append(f"Scryfall error query={q}: {exc}")
            continue
    if not all_cards:
        return None, None, warnings + [f"No se encontró carta para {cleaned}"]
    matched = [c for c in all_cards if _normalized_for_match(c.get("name")) == _normalized_for_match(cleaned)] or all_cards
    if is_foil:
        foil_first = [c for c in matched if c.get("foil")]
        if foil_first:
            matched = foil_first
    if variant_hint:
        for c in matched:
            if variant_hint.lower() in str(c.get("collector_number", "")).lower() or variant_hint.lower() in str(c.get("frame_effects", "")).lower():
                matched = [c]
                break
    unique_names = {c.get("name") for c in matched if c.get("name")}
    if len(matched) > 1 and len(unique_names) > 1:
        sugg = [{"name": c.get("name"), "set_code": c.get("set"), "scryfall_id": c.get("id")} for c in matched[:3]]
        return None, {"suggestions": sugg}, warnings + [f"Ambiguo para {cleaned}"]
    card_data = matched[0]
    card, _ = MTGCard.objects.update_or_create(scryfall_id=card_data["id"], defaults=_normalize_card_data(card_data))
    return card, card_data, warnings

def import_single_catalog_row(row_data):
    card, card_data, warnings = resolve_scryfall_card(scryfall_id=row_data.get("scryfall_id"), name=row_data.get("name"))
    name = str(row_data.get("name") or card.name).strip()
    product, created = Product.objects.update_or_create(
        name=name, product_type=Product.ProductType.SINGLE,
        defaults={"category": row_data.get("category"), "description": str(row_data.get("description") or card.type_line or ""), "price_clp": _to_int(row_data.get("price_clp"), 0), "image": str(row_data.get("image") or card.image_large or card.image_normal or card.image_small or ""), "notes": str(row_data.get("notes") or ""), "is_active": _to_bool(row_data.get("is_active"), True)},
    )
    SingleCard.objects.update_or_create(product=product, defaults={"mtg_card": card, "condition": str(row_data.get("condition") or Product.CardCondition.NM).upper(), "language": str(row_data.get("language") or "EN").upper(), "is_foil": _to_bool(row_data.get("is_foil"), False), "edition": row_data.get("set_name") or card.set_name, "price_usd_reference": extract_usd_price(card_data, _to_bool(row_data.get("is_foil"), False))})
    return product, created, warnings

def import_sealed_catalog_row(row_data):
    sealed_kind = str(row_data.get("sealed_kind") or "").strip().lower()
    if not sealed_kind:
        raise ValidationError("sealed_kind es obligatorio para type=sealed")
    product, created = Product.objects.update_or_create(name=str(row_data.get("name") or "").strip(), product_type=Product.ProductType.SEALED, defaults={"category": row_data.get("category"), "description": str(row_data.get("description") or ""), "price_clp": _to_int(row_data.get("price_clp"), 0), "image": str(row_data.get("image") or ""), "notes": str(row_data.get("notes") or ""), "is_active": _to_bool(row_data.get("is_active"), True)})
    SealedProduct.objects.update_or_create(product=product, defaults={"sealed_kind": sealed_kind, "set_code": str(row_data.get("set_code") or "")})
    return product, created, []

def import_catalog_row(row_data):
    row_type = str(row_data.get("type") or "").strip().lower()
    if not row_type: raise ValidationError("type es obligatorio")
    if not str(row_data.get("name") or "").strip(): raise ValidationError("name es obligatorio")
    price = _to_int(row_data.get("price_clp"), 0)
    if price < 0: raise ValidationError("price_clp debe ser entero >= 0")
    if row_type == Product.ProductType.SINGLE:
        return import_single_catalog_row(row_data)
    if row_type == Product.ProductType.SEALED:
        return import_sealed_catalog_row(row_data)
    if row_type == Product.ProductType.BUNDLE:
        product, created = Product.objects.update_or_create(name=str(row_data.get("name")).strip(), product_type=Product.ProductType.BUNDLE, defaults={"category": row_data.get("category"), "description": str(row_data.get("description") or ""), "price_clp": price, "image": str(row_data.get("image") or ""), "notes": str(row_data.get("notes") or ""), "is_active": _to_bool(row_data.get("is_active"), True)})
        return product, created, []
    raise ValidationError("type inválido. Usa single, sealed o bundle")

def import_catalog_from_xlsx(excel_file):
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook["catalog"] if "catalog" in workbook.sheetnames else workbook.active
    raw_headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized_headers = [_normalize_header(h) for h in raw_headers]
    detected_format = _detect_xlsx_format(normalized_headers)
    header_map = {h: h for h in normalized_headers}
    if detected_format == "catalog":
        _, header_map = _resolve_catalog_headers(raw_headers)
    summary = {"created": 0, "updated": 0, "errors": [], "warnings": [], "preview": [], "detected_format": detected_format}
    categories = {c.name.strip().lower(): c for c in Product._meta.get_field('category').related_model.objects.all()}
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        source_row_data = dict(zip(normalized_headers, row))
        row_data = {canonical: source_row_data.get(source_header) for canonical, source_header in header_map.items()}
        row_data["category"] = categories.get(str(row_data.get("category") or "").strip().lower())
        try:
            logger.info("Procesando fila %s", row_num)
            if detected_format == "single_purchase_items":
                product, created, warns = import_single_purchase_catalog_row(row_data)
            else:
                product, created, warns = import_catalog_row(row_data)
            summary["created" if created else "updated"] += 1
            summary["warnings"].extend([{"row": row_num, "warning": w} for w in warns])
            summary["preview"].append({"row": row_num, "product_id": product.id, "status": "ok"})
        except Exception as exc:
            logger.error("Error fila %s: %s", row_num, exc)
            if isinstance(exc, ValidationError) and hasattr(exc, "message_dict"):
                error_payload = {"row": row_num, **exc.message_dict}
            else:
                error_payload = {"row": row_num, "error": str(exc)}
            summary["errors"].append(error_payload)
            summary["preview"].append({"row": row_num, "status": "error"})
    return summary


def _get_xlsx_sheet(workbook, preferred_sheet=None):
    if preferred_sheet and preferred_sheet in workbook.sheetnames:
        return workbook[preferred_sheet]
    return workbook.active


def _sheet_headers(sheet):
    first_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not first_row:
        raise ValidationError("El archivo XLSX está vacío")
    headers = [str(c.value or "").strip().lower() for c in first_row]
    if not any(headers):
        raise ValidationError("El archivo XLSX no tiene encabezados válidos")
    return headers

def import_purchase_order_from_xlsx(*, excel_file, user, purchase_order_id=None):
    workbook = load_workbook(excel_file, data_only=True)
    sheet = _get_xlsx_sheet(workbook, preferred_sheet="purchase_orders")
    headers = _sheet_headers(sheet)
    detected_format = _detect_xlsx_format(headers)
    if detected_format == "catalog":
        raise ValidationError("Este archivo corresponde a catálogo. Usa /api/products/import-catalog-xlsx/")
    first_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if not first_row:
        raise ValidationError("El XLSX no contiene filas de detalle")
    first = dict(zip(headers, first_row))
    supplier_name = str(first.get("supplier") or "Proveedor XLSX Singles").strip()
    supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
    with transaction.atomic():
        po = PurchaseOrder.objects.filter(pk=purchase_order_id).first() if purchase_order_id else None
        if not po:
            po = PurchaseOrder.objects.create(supplier=supplier, order_number=str(first.get("order_number") or f"XLSX-{timezone.now().timestamp()}"), created_by=user, status=PurchaseOrder.Status.DRAFT, exchange_rate=_to_decimal(first.get("exchange_rate"), Decimal("0")))
        summary = {"rows_processed": 0, "errors": [], "preview": []}
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            r = dict(zip(headers, row)); summary["rows_processed"] += 1
            try:
                qty = _to_int(r.get("quantity") if detected_format == "catalog" else r.get("qty"), 0)
                if qty <= 0: raise ValidationError("quantity/qty debe ser entero > 0")
                product = Product.objects.filter(pk=r.get("product_id")).first()
                if not product and r.get("name"):
                    product = Product.objects.filter(name=str(r.get("name")).strip()).first()
                if not product and detected_format == "single_purchase_items":
                    product, _, _ = import_single_purchase_catalog_row(r)
                if not product: raise ValidationError("No se pudo resolver product_id/name")
                unit_cost_clp = _to_int(r.get("unit_cost_clp"), 0)
                unit_cost_usd = _to_decimal(r.get("unit_cost_usd") if detected_format == "catalog" else r.get("price_usd"), Decimal("0"))
                PurchaseOrderItem.objects.create(purchase_order=po, product=product, quantity_ordered=qty, quantity_received=0, unit_cost_usd=unit_cost_usd, unit_cost_clp=unit_cost_clp, subtotal_clp=qty * unit_cost_clp)
                summary["preview"].append({"row": row_num, "status": "ok", "product_id": product.id})
            except Exception as exc:
                summary["errors"].append({"row": row_num, "error": str(exc)})
                summary["preview"].append({"row": row_num, "status": "error"})
    return po, summary

def calculate_price_clp(usd_price, is_foil=False):
    usd = _to_decimal(usd_price)
    settings = get_active_pricing_settings()
    raw_clp = usd * settings.usd_to_clp
    return {"usd": float(usd), "is_foil": is_foil, "clp_sugerido": int(raw_clp)}


def calculate_suggested_sale_price(product, unit_cost_clp=None):
    unit_cost = int(unit_cost_clp or 0)
    return {"suggested_price_clp": max(unit_cost, int(product.price_clp or 0)), "min_price_clp": unit_cost, "source": "MANUAL"}
