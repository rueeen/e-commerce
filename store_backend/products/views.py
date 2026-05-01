from django.db import transaction
from django.db.models import Count
from django.db import models
from django.core.exceptions import ValidationError
import logging
from openpyxl import load_workbook
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsAdminUser, IsAdminOrWorkerUser
from .models import Category, KardexMovement, MTGCard, PricingSettings, Product, PurchaseOrder, Supplier
from .permissions import IsAdminOrReadOnly
from .serializers import CategorySerializer, KardexMovementSerializer, MTGCardSerializer, PricingSettingsSerializer, ProductSerializer, PurchaseOrderSerializer, SupplierSerializer
from .services import ScryfallServiceError, calculate_price_clp, calculate_suggested_sale_price, extract_usd_price, get_active_pricing_settings, get_scryfall_card_by_id, import_card, search_cards
from .inventory_services import create_stock_movement, receive_purchase_order


logger = logging.getLogger(__name__)


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes", "on"}
    return bool(value)


class CardViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = MTGCard.objects.all()
    serializer_class = MTGCardSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ["name", "set_name", "set_code", "collector_number", "rarity"]


class MTGScryfallViewSet(viewsets.ViewSet):
    def get_permissions(self):
        if self.action == "search":
            return [AllowAny()]
        if self.action == "import_card_action":
            return [IsAdminOrWorkerUser()]
        return [IsAdminUser()]

    @action(detail=False, methods=["get"], url_path="search")
    def search(self, request):
        query = request.query_params.get("q", "").strip()
        if not query:
            return Response({"detail": "q es obligatorio"}, status=400)
        try:
            return Response({"results": search_cards(query)})
        except ScryfallServiceError as exc:
            return Response({"detail": str(exc)}, status=502)

    @action(detail=False, methods=["post"], url_path="import")
    def import_card_action(self, request):
        scryfall_id = request.data.get("scryfall_id")
        if not scryfall_id:
            return Response({"detail": "scryfall_id es obligatorio"}, status=400)
        try:
            card = import_card(scryfall_id)
        except ScryfallServiceError as exc:
            return Response({"detail": str(exc)}, status=502)
        return Response(MTGCardSerializer(card).data, status=201)


class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name", "description", "mtg_card__name"]
    ordering_fields = ["price", "price_clp", "created_at", "name", "stock"]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_queryset(self):
        q = Product.objects.select_related("mtg_card", "category").all()
        p = self.request.query_params
        if p.get("product_type"):
            q = q.filter(product_type=p["product_type"])
        if p.get("category"):
            q = q.filter(category_id=p["category"])
        if p.get("active") in {"true", "false"}:
            q = q.filter(is_active=p["active"] == "true")
        if p.get("rarity"):
            q = q.filter(mtg_card__rarity__iexact=p["rarity"])
        return q

    @action(detail=False, methods=["post"], url_path="create-single-from-scryfall", permission_classes=[IsAdminUser])
    def create_single_from_scryfall(self, request):
        required_fields = ["scryfall_id", "category_id", "price_clp_final", "condition", "language"]
        errors = {}
        payload = request.data or {}
        for field in required_fields:
            if payload.get(field) in (None, ""):
                errors[field] = "Este campo es obligatorio."
        if errors:
            return Response({"detail": "Payload inválido", "errors": errors}, status=400)

        try:
            category_id = int(payload.get("category_id"))
            price_clp_final = int(payload.get("price_clp_final", 0))
            is_foil = _to_bool(payload.get("is_foil", False))
            is_active = _to_bool(payload.get("is_active", True), default=True)
            condition = str(payload.get("condition", "")).strip()
            language = str(payload.get("language", "")).strip().upper()
            notes = str(payload.get("notes", "") or "").strip()
            scryfall_id = str(payload.get("scryfall_id", "")).strip()
        except (TypeError, ValueError):
            return Response({"detail": "category_id y price_clp_final deben ser numéricos válidos"}, status=400)

        category = Category.objects.filter(pk=category_id).first()
        if not category:
            return Response({"detail": "category_id inválido", "errors": {"category_id": "No existe la categoría indicada."}}, status=400)

        if not scryfall_id:
            return Response({"detail": "Payload inválido", "errors": {"scryfall_id": "Este campo es obligatorio."}}, status=400)
        if not condition:
            return Response({"detail": "Payload inválido", "errors": {"condition": "Este campo es obligatorio."}}, status=400)
        if not language:
            return Response({"detail": "Payload inválido", "errors": {"language": "Este campo es obligatorio."}}, status=400)
        if price_clp_final < 0:
            return Response({"detail": "price_clp_final no puede ser menor a 0", "errors": {"price_clp_final": "Debe ser mayor o igual a 0."}}, status=400)

        try:
            card_data = get_scryfall_card_by_id(scryfall_id)
            image_uris = card_data.get("image_uris") or {}
            faces = card_data.get("card_faces") or []
            image_large = image_uris.get("large") or image_uris.get("normal")
            if not image_large and faces and faces[0].get("image_uris"):
                image_large = (
                    faces[0]["image_uris"].get("large")
                    or faces[0]["image_uris"].get("normal")
                )

            image_normal = image_uris.get("normal")
            if not image_normal and faces and faces[0].get("image_uris"):
                image_normal = faces[0]["image_uris"].get("normal")

            image_small = image_uris.get("small")
            if not image_small and faces and faces[0].get("image_uris"):
                image_small = faces[0]["image_uris"].get("small")

            card_defaults = {
                "name": card_data.get("name", ""),
                "set_code": card_data.get("set", ""),
                "set_name": card_data.get("set_name", ""),
                "collector_number": card_data.get("collector_number", ""),
                "rarity": card_data.get("rarity", ""),
                "type_line": card_data.get("type_line", ""),
                "oracle_text": card_data.get("oracle_text", ""),
                "image_large": image_large or "",
                "raw_data": card_data,
            }
            if hasattr(MTGCard, "image_normal"):
                card_defaults["image_normal"] = image_normal or ""
            if hasattr(MTGCard, "image_small"):
                card_defaults["image_small"] = image_small or ""

            card, _ = MTGCard.objects.update_or_create(
                scryfall_id=card_data["id"],
                defaults=card_defaults,
            )

            usd_ref = extract_usd_price(card_data, is_foil=is_foil)
            pricing = calculate_price_clp(usd_ref, is_foil=is_foil)
            if price_clp_final == 0 and pricing.get("clp_sugerido"):
                price_clp_final = int(pricing["clp_sugerido"])

            with transaction.atomic():
                product = Product.objects.create(
                    mtg_card=card,
                    category=category,
                    name=f"{card.name} - {card.set_code.upper()} #{card.collector_number}",
                    description=f"{card.type_line}\nRareza: {card.rarity}\nSet: {card.set_name} ({card.set_code.upper()})\n\n{card.oracle_text}",
                    price_clp=price_clp_final,
                    price=price_clp_final,
                    price_usd_reference=usd_ref,
                    price_clp_suggested=pricing["clp_sugerido"],
                    price_clp_final=price_clp_final,
                    stock=0,
                    condition=condition,
                    language=language,
                    is_foil=is_foil,
                    notes=notes,
                    is_active=is_active,
                    product_type=Product.ProductType.SINGLE,
                    image=card.image_large or card.image_normal or card.image_small,
                )
        except ValidationError as exc:
            url = f"https://api.scryfall.com/cards/{scryfall_id}"
            logger.error("Error consultando Scryfall URL=%s status=400 response=%s", url, exc, exc_info=True)
            return Response({
                "detail": "No se pudo obtener la carta desde Scryfall usando el ID recibido.",
                "scryfall_id": scryfall_id,
                "scryfall_response": str(exc),
            }, status=400)
        except ScryfallServiceError as exc:
            return Response({
                "detail": "No se pudo obtener la carta desde Scryfall.",
                "scryfall_id": scryfall_id,
                "scryfall_response": str(exc),
            }, status=400)

        return Response({
            "id": product.id,
            "name": product.name,
            "price_clp_final": product.price_clp_final,
            "stock": product.stock,
            "mtg_card": product.mtg_card_id,
            "image": product.image,
            "category": product.category_id,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"], url_path="suggested-price", permission_classes=[IsAdminUser])
    def suggested_price(self, request, pk=None):
        product = self.get_object()
        unit_cost_clp = request.query_params.get("unit_cost_clp", 0)
        try:
            unit_cost_clp = int(float(unit_cost_clp or 0))
        except (TypeError, ValueError):
            return Response({"detail": "unit_cost_clp inválido"}, status=400)
        return Response(calculate_suggested_sale_price(product, unit_cost_clp=unit_cost_clp))

    @action(detail=True, methods=["get"], url_path="kardex", permission_classes=[IsAdminOrWorkerUser])
    def kardex(self, request, pk=None):
        product = self.get_object()
        movements = product.kardex_movements.all()[:50]
        return Response(KardexMovementSerializer(movements, many=True).data)


class KardexViewSet(viewsets.GenericViewSet):
    serializer_class = KardexMovementSerializer
    permission_classes = [IsAdminOrWorkerUser]

    def get_queryset(self):
        qs = KardexMovement.objects.select_related("product", "created_by")
        product_id = self.request.query_params.get("product_id")
        if product_id:
            qs = qs.filter(product_id=product_id)
        mt = self.request.query_params.get("movement_type")
        if mt:
            qs = qs.filter(movement_type=mt)
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        supplier_id = self.request.query_params.get("supplier_id")
        if supplier_id:
            qs = qs.filter(reference_type="PURCHASE_ORDER", reference_id__in=PurchaseOrder.objects.filter(supplier_id=supplier_id).values_list("id", flat=True))
        return qs

    def list(self, request):
        return Response(self.get_serializer(self.get_queryset()[:200], many=True).data)

    @action(detail=False, methods=["post"], url_path="movement")
    def movement(self, request):
        p = request.data
        product = Product.objects.filter(pk=p.get("product")).first()
        if not product:
            return Response({"detail": "product inválido"}, status=400)
        try:
            qty = int(p.get("quantity", 0))
            if qty <= 0:
                return Response({"detail": "quantity debe ser mayor a 0"}, status=400)
            movement = create_stock_movement(product=product, movement_type=p.get("movement_type"), quantity=qty, created_by=request.user, unit_cost_clp=int(p.get("unit_cost_clp", 0) or 0), unit_price_clp=int(p.get("unit_price_clp", 0) or 0), reference_label=p.get("reference_label", ""), reference_type=p.get("reference_type", "manual"), reference_id=p.get("reference_id", ""), notes=p.get("notes", ""))
        except (TypeError, ValueError, ValidationError) as exc:
            return Response({"detail": str(exc)}, status=400)
        return Response(self.get_serializer(movement).data, status=201)

    @action(detail=False, methods=["post"], url_path="import-excel")
    def import_excel(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"detail": "Debes adjuntar un archivo .xlsx"}, status=status.HTTP_400_BAD_REQUEST)
        workbook = load_workbook(excel_file, data_only=True)
        sheet = workbook.active
        headers = [str(c.value or "").strip().lower() for c in next(sheet.iter_rows(min_row=1, max_row=1))]
        required = {"category", "product_type", "name", "price_clp", "stock", "is_active"}
        if not required.issubset(set(headers)):
            return Response({"detail": "Columnas inválidas", "required": sorted(required)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({"detail": "Importación Excel no modificada en esta tarea"})


class CategoryViewSet(viewsets.ModelViewSet):
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ["name", "slug"]

    def get_queryset(self):
        return Category.objects.annotate(products_count=Count("products")).order_by("name")


class PricingSettingsViewSet(viewsets.ModelViewSet):
    serializer_class = PricingSettingsSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        return PricingSettings.objects.order_by("-updated_at")


    def _ensure_single_active(self, instance):
        if instance.is_active:
            PricingSettings.objects.exclude(pk=instance.pk).update(is_active=False)

    def perform_create(self, serializer):
        instance = serializer.save()
        self._ensure_single_active(instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        self._ensure_single_active(instance)

    @action(detail=False, methods=["get"], url_path="active", permission_classes=[AllowAny])
    def active(self, request):
        active_settings = get_active_pricing_settings()
        return Response({"usd_to_clp": active_settings.usd_to_clp, "import_factor": active_settings.import_factor, "risk_factor": active_settings.risk_factor, "margin_factor": active_settings.margin_factor, "rounding_to": active_settings.rounding_to})


class SupplierViewSet(viewsets.ModelViewSet):
    serializer_class = SupplierSerializer
    permission_classes = [IsAdminOrWorkerUser]

    def get_queryset(self):
        return Supplier.objects.order_by("name")


class PurchaseOrderViewSet(viewsets.ModelViewSet):
    serializer_class = PurchaseOrderSerializer
    permission_classes = [IsAdminOrWorkerUser]

    def get_queryset(self):
        return PurchaseOrder.objects.select_related("supplier", "created_by").prefetch_related("items__product").order_by("-created_at")

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"], url_path="receive")
    def receive(self, request, pk=None):
        po = receive_purchase_order(int(pk), request.user)
        return Response(self.get_serializer(po).data)


class InventoryDashboardView(APIView):
    permission_classes = [IsAdminOrWorkerUser]

    def get(self, request):
        products = Product.objects.all()
        inventory_value_avg_cost = sum(int(p.stock or 0) * int(p.average_cost_clp or 0) for p in products)
        out_of_stock = Product.objects.filter(stock=0)[:50]
        low_stock = Product.objects.filter(stock__lte=models.F("stock_minimum")).exclude(stock_minimum=0)[:50]
        latest_in = KardexMovement.objects.filter(movement_type=KardexMovement.MovementType.PURCHASE_IN)[:10]
        latest_out = KardexMovement.objects.filter(movement_type=KardexMovement.MovementType.SALE_OUT)[:10]
        pending_po = PurchaseOrder.objects.filter(status__in=[PurchaseOrder.Status.DRAFT, PurchaseOrder.Status.SENT]).order_by("-created_at")[:20]
        return Response({
            "inventory_value_avg_cost_clp": inventory_value_avg_cost,
            "products_without_stock": ProductSerializer(out_of_stock, many=True).data,
            "products_below_minimum_stock": ProductSerializer(low_stock, many=True).data,
            "latest_entries": KardexMovementSerializer(latest_in, many=True).data,
            "latest_exits": KardexMovementSerializer(latest_out, many=True).data,
            "purchase_orders_pending_receipt": PurchaseOrderSerializer(pending_po, many=True).data,
        })
